from PIL import Image
from openpyxl import load_workbook
import time

def start_conversion(file):
    # Open the Excel workbook
    workbook = load_workbook(file, data_only=True)

    # Read in the first sheet from Excel workbook
    sheet = workbook.worksheets[0]

    # Get the total number of columns in the sheet
    cols = sheet.iter_cols()

    # Create a new, blank image the size of the worksheet
    image = Image.new("RGB", (sheet.max_row, sheet.max_column))
    pixels = image.load()

    # Iterate over all the columns and rows, get the background color
    # and write it to the image
    for y, col in enumerate(cols):
        for x, cell in enumerate(col):
            index = cell.fill.start_color.index
            rgb = hex_to_rgb("#" + str(index[2:]))
            pixels[x, y] = rgb

    # Rotate the image clockwise 90 degrees
    image = image.rotate(-90)

    # Save the new image
    image.save("output.jpg")
            

# Convert hex to (r, g, b)
def hex_to_rgb(vhex):
    strip = vhex.lstrip('#')
    return tuple(int(strip[i:i+2], 16) for i in (0, 2, 4))


if __name__ == '__main__':
    # Get the file name from user
    file = input("Enter XLS file name: ")

    # Let user know it's working
    print("Converting your Excel file to an image...")

    # Start timer 
    start_time = time.time()

    # Convert image to Excel cells
    start_conversion(file)

    # Print out how long it took
    end_time = time.time()
    print("Time taken: {}".format(end_time - start_time))
